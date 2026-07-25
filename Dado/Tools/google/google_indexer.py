"""Build Dado's private, read-only Google reference index.

The index is stored outside the FRP Depot Git repository under
%LOCALAPPDATA%\FRPDepot-Google\reference. Google is never modified.
Gmail remains screened. Drive is unrestricted by Rachad's instruction.
"""
from __future__ import annotations

import argparse
import base64
import csv
import hashlib
import io
import json
import os
import re
import sqlite3
import sys
import time
import zipfile
from datetime import datetime, timezone
from email.utils import parseaddr
from html import unescape
from pathlib import Path
from typing import Any, Iterable
from xml.etree import ElementTree as ET

from bs4 import BeautifulSoup
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload
from openpyxl import load_workbook
from pypdf import PdfReader
from docx import Document

sys.path.insert(0, str(Path(__file__).resolve().parent))
import google_auth
from tdi_filter import deep_tdi_marker

ROOT = Path(r"C:\FRPDepot")
VAULT = Path(os.environ["LOCALAPPDATA"]) / "FRPDepot-Google" / "reference"
DB_PATH = VAULT / "google_reference.sqlite"
REPORT_PATH = VAULT / "google_index_report.md"
RECEIPTS = ROOT / "Dado" / "40_Logs" / "receipts.jsonl"
BATCH_SIZE = 20
MAX_TEXT_PER_FILE = 500_000
MAX_DOWNLOAD_BYTES = 30 * 1024 * 1024
LIST_PAGE_SIZE = 1000       # Drive/Gmail allow far more than the 100 first used
FOLDER_MIME = "application/vnd.google-apps.folder"

# A run must always end on its own. The 2026-07-24 run was unbounded: Rachad's
# Drive holds 114,000+ items, it had done 44,400 in three hours, and it would
# have needed ~16 more. An unattended job with no ceiling is a runaway.
DEFAULT_MAX_MINUTES = 45


def now() -> str:
    return datetime.now(timezone.utc).isoformat()


def receipt(action: str, evidence: str) -> None:
    RECEIPTS.parent.mkdir(parents=True, exist_ok=True)
    with RECEIPTS.open("a", encoding="utf-8") as fh:
        fh.write(json.dumps({"ts": now(), "action": action, "evidence": evidence}, ensure_ascii=True) + "\n")


def chunks(items: list[Any], n: int = BATCH_SIZE) -> Iterable[list[Any]]:
    for i in range(0, len(items), n):
        yield items[i:i+n]


class Budget:
    """Wall-clock ceiling so an unattended run always stops and can resume."""

    def __init__(self, minutes: float | None) -> None:
        self.deadline = None if not minutes else time.monotonic() + minutes * 60

    def spent(self) -> bool:
        return self.deadline is not None and time.monotonic() >= self.deadline


def sha(value: str) -> str:
    return hashlib.sha256(value.encode("utf-8", "replace")).hexdigest()


def db_open() -> sqlite3.Connection:
    VAULT.mkdir(parents=True, exist_ok=True)
    con = sqlite3.connect(DB_PATH)
    con.execute("PRAGMA journal_mode=WAL")
    con.execute("PRAGMA synchronous=NORMAL")
    con.executescript("""
    CREATE TABLE IF NOT EXISTS gmail_messages (
      id TEXT PRIMARY KEY, thread_id TEXT, internal_ts INTEGER, date_header TEXT,
      subject TEXT, sender TEXT, recipients TEXT, cc TEXT, labels_json TEXT,
      snippet TEXT, body TEXT, attachment_names TEXT, indexed_at TEXT
    );
    CREATE VIRTUAL TABLE IF NOT EXISTS gmail_fts USING fts5(
      id UNINDEXED, subject, sender, recipients, cc, snippet, body, attachment_names
    );
    CREATE TABLE IF NOT EXISTS drive_files (
      id TEXT PRIMARY KEY, name TEXT, mime_type TEXT, created_time TEXT,
      modified_time TEXT, size INTEGER, owners TEXT, parents_json TEXT,
      web_view_link TEXT, description TEXT, content TEXT, content_status TEXT,
      indexed_at TEXT
    );
    CREATE VIRTUAL TABLE IF NOT EXISTS drive_fts USING fts5(
      id UNINDEXED, name, mime_type, owners, description, content
    );
    CREATE TABLE IF NOT EXISTS withheld_hashes (
      kind TEXT NOT NULL, id_hash TEXT NOT NULL, screened_at TEXT NOT NULL,
      PRIMARY KEY(kind, id_hash)
    );
    CREATE TABLE IF NOT EXISTS meta (key TEXT PRIMARY KEY, value TEXT);
    """)
    # Quarantine columns remain for Gmail and schema compatibility. Drive does
    # not use them and every Drive row is queryable.
    for table in ("gmail_messages", "drive_files"):
        cols = {r[1] for r in con.execute(f"PRAGMA table_info({table})")}
        if "tdi_quarantined" not in cols:
            con.execute(f"ALTER TABLE {table} ADD COLUMN tdi_quarantined INTEGER DEFAULT 0")
        if "tdi_marker" not in cols:
            con.execute(f"ALTER TABLE {table} ADD COLUMN tdi_marker TEXT")
    con.execute("INSERT OR REPLACE INTO meta VALUES('rescreen_marker_version','2026-07-24-deep')")
    con.commit()
    return con


def b64decode(data: str) -> bytes:
    return base64.urlsafe_b64decode(data + "=" * (-len(data) % 4))


def text_from_payload(payload: dict[str, Any]) -> tuple[str, list[str]]:
    plain: list[str] = []
    html: list[str] = []
    attachments: list[str] = []

    def walk(part: dict[str, Any]) -> None:
        filename = part.get("filename") or ""
        if filename:
            attachments.append(filename)
        mime = (part.get("mimeType") or "").lower()
        data = (part.get("body") or {}).get("data")
        if data and mime in {"text/plain", "text/html"}:
            try:
                decoded = b64decode(data).decode("utf-8", "replace")
                (html if mime == "text/html" else plain).append(decoded)
            except Exception:
                pass
        for child in part.get("parts") or []:
            walk(child)

    walk(payload or {})
    if plain:
        text = "\n".join(plain)
    elif html:
        text = "\n".join(BeautifulSoup(x, "html.parser").get_text(" ", strip=True) for x in html)
    else:
        text = ""
    return text, attachments


def headers_dict(msg: dict[str, Any]) -> tuple[dict[str, str], list[str]]:
    values: list[str] = []
    selected: dict[str, str] = {}
    wanted = {"subject", "from", "to", "cc", "date"}
    for item in (msg.get("payload") or {}).get("headers") or []:
        name = (item.get("name") or "").lower()
        value = item.get("value") or ""
        values.append(value)
        if name in wanted and name not in selected:
            selected[name] = value
    return selected, values


def fts_clear(con: sqlite3.Connection, table: str, row_id: str, known_new: bool) -> None:
    """Drop any stale FTS row for this id -- but ONLY when one can exist.

    'id' is UNINDEXED in these FTS5 tables, so "DELETE FROM <fts> WHERE id=?"
    cannot use an index and degenerates into a full scan of the whole FTS table.
    The indexer only ever processes ids that are NOT already stored, so on the
    hot path that scan is guaranteed to delete nothing -- pure waste that grows
    with the table and turns the run quadratic. (Measured 2026-07-24: throughput
    fell from 255 to 73 items/min as the table grew past 40k rows.) Skipping it
    for known-new ids keeps the re-index path correct while making the common
    path O(1).
    """
    if known_new:
        return
    con.execute(f"DELETE FROM {table} WHERE id=?", (row_id,))


def upsert_gmail(con: sqlite3.Connection, msg: dict[str, Any], label_names: dict[str, str],
                 known_new: bool = False) -> str:
    h, all_header_values = headers_dict(msg)
    body, attachment_names = text_from_payload(msg.get("payload") or {})
    labels = [label_names.get(x, x) for x in msg.get("labelIds") or []]
    screen_fields = all_header_values + [msg.get("snippet") or "", body] + attachment_names + labels
    if deep_tdi_marker(*screen_fields):
        con.execute("INSERT OR REPLACE INTO withheld_hashes(kind,id_hash,screened_at) VALUES('gmail',?,?)",
                    (sha(msg["id"]), now()))
        return "withheld"
    values = (
        msg["id"], msg.get("threadId") or "", int(msg.get("internalDate") or 0), h.get("date", ""),
        h.get("subject", ""), h.get("from", ""), h.get("to", ""), h.get("cc", ""),
        json.dumps(labels, ensure_ascii=False), msg.get("snippet") or "", body,
        " | ".join(attachment_names), now()
    )
    fts_clear(con, "gmail_fts", msg["id"], known_new)
    # Columns are named, not positional: the quarantine columns were added later
    # and a bare VALUES(?,...) silently became an arity error on every insert.
    con.execute(
        "INSERT OR REPLACE INTO gmail_messages"
        " (id,thread_id,internal_ts,date_header,subject,sender,recipients,cc,"
        "  labels_json,snippet,body,attachment_names,indexed_at)"
        " VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)", values)
    con.execute("INSERT INTO gmail_fts VALUES(?,?,?,?,?,?,?,?)",
                (msg["id"], h.get("subject", ""), h.get("from", ""), h.get("to", ""),
                 h.get("cc", ""), msg.get("snippet") or "", body, " | ".join(attachment_names)))
    return "indexed"


def gmail_index(con: sqlite3.Connection, gmail: Any, max_items: int | None,
                budget: Budget) -> dict[str, int]:
    profile = gmail.users().getProfile(userId="me").execute(num_retries=3)
    if profile.get("emailAddress", "").lower() != "rachad85@gmail.com":
        raise RuntimeError("Connected Gmail account is not rachad85@gmail.com; stopped without indexing")
    con.execute("INSERT OR REPLACE INTO meta VALUES('gmail_sweep_status','running')")
    con.commit()
    labels_data = gmail.users().labels().list(userId="me").execute(num_retries=3).get("labels", [])
    labels = {x["id"]: x.get("name", x["id"]) for x in labels_data}
    stats = {"listed": 0, "fetched": 0, "indexed": 0, "withheld": 0, "errors": 0,
             "already_indexed": 0, "stopped_early": 0}
    page_token = None
    stop = False
    while not stop:
        listing = gmail.users().messages().list(userId="me", maxResults=500, pageToken=page_token).execute(num_retries=3)
        ids = [x["id"] for x in listing.get("messages", [])]
        stats["listed"] += len(ids)
        pending: list[str] = []
        for mid in ids:
            if con.execute("SELECT 1 FROM gmail_messages WHERE id=?", (mid,)).fetchone():
                stats["already_indexed"] += 1
                continue
            if con.execute("SELECT 1 FROM withheld_hashes WHERE kind='gmail' AND id_hash=?", (sha(mid),)).fetchone():
                stats["already_indexed"] += 1
                continue
            pending.append(mid)
        for group in chunks(pending):
            if max_items is not None:
                remaining = max_items - stats["fetched"]
                if remaining <= 0:
                    stop = True
                    break
                group = group[:remaining]
            responses: dict[str, dict[str, Any]] = {}
            failures: set[str] = set()

            def callback(request_id: str, response: Any, exception: Exception | None) -> None:
                if exception is not None or not response:
                    failures.add(request_id)
                else:
                    responses[request_id] = response

            batch = gmail.new_batch_http_request(callback=callback)
            for mid in group:
                batch.add(gmail.users().messages().get(userId="me", id=mid, format="full"), request_id=mid)
            # BatchHttpRequest has no num_retries parameter in this installed
            # Google client version. Individual failed responses are captured by
            # the callback and counted without stopping the checkpoint.
            batch.execute()
            for mid in group:
                stats["fetched"] += 1
                if mid in failures or mid not in responses:
                    stats["errors"] += 1
                    continue
                result = upsert_gmail(con, responses[mid], labels, known_new=True)
                stats[result] += 1
            con.commit()
            print(json.dumps({"phase": "gmail", **stats}), flush=True)
            if budget.spent():
                stats["stopped_early"] = 1
                stop = True
                break
            if max_items is not None and stats["fetched"] >= max_items:
                stop = True
                break
        if stop:
            break
        page_token = listing.get("nextPageToken")
        if not page_token:
            break
    if stats["stopped_early"]:
        status = "incomplete_time_limit"
    elif stats["errors"]:
        status = f"complete_with_{stats['errors']}_fetch_errors"
    else:
        status = "complete"
    con.execute("INSERT OR REPLACE INTO meta VALUES('gmail_sweep_status',?)", (status,))
    con.execute("INSERT OR REPLACE INTO meta VALUES('gmail_last_sweep',?)", (now(),))
    con.commit()
    return stats


def clean_html(raw: bytes) -> str:
    return BeautifulSoup(raw.decode("utf-8", "replace"), "html.parser").get_text(" ", strip=True)


def parse_pptx(raw: bytes) -> str:
    out: list[str] = []
    with zipfile.ZipFile(io.BytesIO(raw)) as zf:
        for name in sorted(x for x in zf.namelist() if re.fullmatch(r"ppt/slides/slide\d+\.xml", x)):
            root = ET.fromstring(zf.read(name))
            out.extend(node.text or "" for node in root.iter() if node.tag.endswith("}t"))
    return "\n".join(out)


def parse_xlsx(raw: bytes) -> str:
    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    out: list[str] = []
    for ws in wb.worksheets:
        out.append(f"[Sheet: {ws.title}]")
        for row in ws.iter_rows(values_only=True):
            line = "\t".join("" if v is None else str(v) for v in row)
            if line.strip():
                out.append(line)
            if sum(len(x) + 1 for x in out) > MAX_TEXT_PER_FILE:
                break
        if sum(len(x) + 1 for x in out) > MAX_TEXT_PER_FILE:
            break
    return "\n".join(out)


def parse_docx(raw: bytes) -> str:
    doc = Document(io.BytesIO(raw))
    return "\n".join(p.text for p in doc.paragraphs)


def parse_pdf(raw: bytes) -> str:
    reader = PdfReader(io.BytesIO(raw))
    out: list[str] = []
    for page in reader.pages:
        out.append(page.extract_text() or "")
        if sum(len(x) for x in out) > MAX_TEXT_PER_FILE:
            break
    return "\n".join(out)


def download_bytes(request: Any) -> bytes:
    fh = io.BytesIO()
    downloader = MediaIoBaseDownload(fh, request, chunksize=4 * 1024 * 1024)
    done = False
    while not done:
        status, done = downloader.next_chunk()
        if fh.tell() > MAX_DOWNLOAD_BYTES:
            raise ValueError("file exceeds local indexing download limit")
    return fh.getvalue()


def extract_drive_content(drive: Any, item: dict[str, Any]) -> tuple[str, str]:
    mime = item.get("mimeType") or ""
    size = int(item.get("size") or 0)
    if mime == "application/vnd.google-apps.folder":
        return "", "folder_metadata"
    if size and size > MAX_DOWNLOAD_BYTES:
        return "", "metadata_only_too_large"
    export_mime = None
    parser = None
    if mime == "application/vnd.google-apps.document":
        export_mime, parser = "text/plain", lambda b: b.decode("utf-8", "replace")
    elif mime == "application/vnd.google-apps.spreadsheet":
        export_mime, parser = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", parse_xlsx
    elif mime == "application/vnd.google-apps.presentation":
        export_mime, parser = "application/vnd.openxmlformats-officedocument.presentationml.presentation", parse_pptx
    elif mime == "application/vnd.google-apps.drawing":
        export_mime, parser = "application/pdf", parse_pdf
    elif mime.startswith("text/") or mime in {"application/json", "application/xml", "application/javascript"}:
        parser = lambda b: b.decode("utf-8", "replace")
    elif mime == "application/pdf":
        parser = parse_pdf
    elif mime == "application/vnd.openxmlformats-officedocument.wordprocessingml.document":
        parser = parse_docx
    elif mime == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
        parser = parse_xlsx
    elif mime == "application/vnd.openxmlformats-officedocument.presentationml.presentation":
        parser = parse_pptx
    elif mime in {"text/html", "application/xhtml+xml"}:
        parser = clean_html
    else:
        return "", "metadata_only_unreadable_type"
    request = drive.files().export_media(fileId=item["id"], mimeType=export_mime) if export_mime else drive.files().get_media(fileId=item["id"])
    raw = download_bytes(request)
    text = unescape(parser(raw))
    truncated = len(text) > MAX_TEXT_PER_FILE
    return text[:MAX_TEXT_PER_FILE], "indexed_truncated" if truncated else "indexed"


def upsert_drive(con: sqlite3.Connection, drive: Any, item: dict[str, Any],
                 known_new: bool = False) -> str:
    owners = " | ".join((x.get("displayName") or x.get("emailAddress") or "") for x in item.get("owners") or [])
    name = item.get("name") or ""
    try:
        content, status = extract_drive_content(drive, item)
    except Exception as exc:
        content, status = "", f"extract_error:{type(exc).__name__}"
    values = (
        item["id"], item.get("name") or "", item.get("mimeType") or "", item.get("createdTime") or "",
        item.get("modifiedTime") or "", int(item.get("size") or 0), owners,
        json.dumps(item.get("parents") or []), item.get("webViewLink") or "", item.get("description") or "",
        content, status, now()
    )
    fts_clear(con, "drive_fts", item["id"], known_new)
    con.execute(
        "INSERT OR REPLACE INTO drive_files"
        " (id,name,mime_type,created_time,modified_time,size,owners,parents_json,"
        "  web_view_link,description,content,content_status,indexed_at)"
        " VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)", values)
    con.execute("INSERT INTO drive_fts VALUES(?,?,?,?,?,?)",
                (item["id"], item.get("name") or "", item.get("mimeType") or "", owners,
                 item.get("description") or "", content))
    return "indexed" if status.startswith("indexed") else "metadata_only"


def drive_index(con: sqlite3.Connection, drive: Any, max_items: int | None,
                budget: Budget) -> dict[str, int]:
    about = drive.about().get(fields="user").execute(num_retries=3).get("user", {})
    if about.get("emailAddress", "").lower() != "rachad85@gmail.com":
        raise RuntimeError("Connected Drive account is not rachad85@gmail.com; stopped without indexing")
    con.execute("INSERT OR REPLACE INTO meta VALUES('drive_sweep_status','running')")
    con.commit()
    stats = {"listed": 0, "processed": 0, "indexed": 0, "metadata_only": 0, "withheld": 0,
             "errors": 0, "already_indexed": 0, "stopped_early": 0}
    page_token = None
    stop = False
    fields = "nextPageToken,files(id,name,mimeType,createdTime,modifiedTime,size,owners(displayName,emailAddress),parents,webViewLink,description,trashed)"
    # Folders carry no content and were 73% of everything walked on 2026-07-24
    # (72,943 of 114,000+ items; 35,302 of the 40,763 rows stored held nothing
    # but a name). A reference index of documents has no use for them.
    query = f"trashed=false and mimeType != '{FOLDER_MIME}'"
    while not stop:
        listing = drive.files().list(pageSize=LIST_PAGE_SIZE, pageToken=page_token,
                                     q=query, fields=fields).execute(num_retries=3)
        items = listing.get("files", [])
        stats["listed"] += len(items)
        pending: list[dict[str, Any]] = []
        for item in items:
            fid = item["id"]
            if con.execute("SELECT 1 FROM drive_files WHERE id=?", (fid,)).fetchone():
                stats["already_indexed"] += 1
                continue
            pending.append(item)
        for group in chunks(pending):
            if max_items is not None:
                remaining = max_items - stats["processed"]
                if remaining <= 0:
                    stop = True
                    break
                group = group[:remaining]
            for item in group:
                try:
                    result = upsert_drive(con, drive, item, known_new=True)
                    stats[result] += 1
                except Exception as exc:
                    # Keep the FIRST error text. A bare counter hid a schema
                    # arity break that was failing EVERY insert and still just
                    # read as "errors: 2" -- indistinguishable from two bad PDFs.
                    stats["errors"] += 1
                    stats.setdefault("first_error", f"{type(exc).__name__}: {exc}"[:200])
                stats["processed"] += 1
            con.commit()
            # Progress goes to stdout only. It used to also append a receipt per
            # 20 items, which wrote 3,768 lines in one day and buried the real
            # business receipts the SOUL rule exists to preserve.
            print(json.dumps({"phase": "drive", **stats}), flush=True)
            if budget.spent():
                stats["stopped_early"] = 1
                stop = True
                break
            if max_items is not None and stats["processed"] >= max_items:
                stop = True
                break
        if stop:
            break
        page_token = listing.get("nextPageToken")
        if not page_token:
            break
    if stats["stopped_early"]:
        status = "incomplete_time_limit"
    elif stats["errors"]:
        status = f"complete_with_{stats['errors']}_processing_errors"
    else:
        status = "complete"
    con.execute("INSERT OR REPLACE INTO meta VALUES('drive_sweep_status',?)", (status,))
    con.execute("INSERT OR REPLACE INTO meta VALUES('drive_last_sweep',?)", (now(),))
    con.commit()
    return stats


def report(con: sqlite3.Connection, gmail_stats: dict[str, int] | None, drive_stats: dict[str, int] | None) -> None:
    gm_count = con.execute("SELECT count(*) FROM gmail_messages").fetchone()[0]
    dr_count = con.execute("SELECT count(*) FROM drive_files").fetchone()[0]
    gm_dates = con.execute("SELECT min(internal_ts),max(internal_ts) FROM gmail_messages").fetchone()
    def fmt_ms(v: int | None) -> str:
        return datetime.fromtimestamp(v / 1000, tz=timezone.utc).date().isoformat() if v else "n/a"
    top_domains: dict[str, int] = {}
    for (sender,) in con.execute("SELECT sender FROM gmail_messages"):
        addr = parseaddr(sender or "")[1].lower()
        domain = addr.rsplit("@", 1)[-1] if "@" in addr else "(unknown)"
        top_domains[domain] = top_domains.get(domain, 0) + 1
    top = sorted(top_domains.items(), key=lambda x: (-x[1], x[0]))[:25]
    type_rows = con.execute("SELECT mime_type,count(*) FROM drive_files GROUP BY mime_type ORDER BY count(*) DESC LIMIT 25").fetchall()
    status_rows = con.execute("SELECT content_status,count(*) FROM drive_files GROUP BY content_status ORDER BY count(*) DESC").fetchall()
    withheld_gm = con.execute("SELECT count(*) FROM withheld_hashes WHERE kind='gmail'").fetchone()[0]

    meta = dict(con.execute("SELECT key,value FROM meta").fetchall())
    lines = [
        "# Private Google reference index", "", f"Updated: {now()}", "",
        "Stored outside the FRP Depot Git repository. Read-only source access. Gmail is screened; Drive is unrestricted.", "",
        "## Coverage", "",
        f"- Gmail sweep status: {meta.get('gmail_sweep_status', 'unknown')}",
        f"- Gmail messages indexed: {gm_count:,}", f"- Gmail date range: {fmt_ms(gm_dates[0])} to {fmt_ms(gm_dates[1])}",
        f"- Gmail TDI-flagged messages withheld: {withheld_gm:,}",
        f"- Drive sweep status: {meta.get('drive_sweep_status', 'unknown')}",
        f"- Drive files indexed/listed: {dr_count:,}",
        "- Drive screening: none (Rachad's instruction)", "",
        "## Gmail top sender domains", "",
    ]
    lines.extend(f"- {domain}: {count:,}" for domain, count in top)
    lines += ["", "## Drive MIME types", ""]
    lines.extend(f"- {mime}: {count:,}" for mime, count in type_rows)
    lines += ["", "## Drive content status", ""]
    lines.extend(f"- {status}: {count:,}" for status, count in status_rows)
    lines += ["", "## Honest limits", "",
              "- Gmail message bodies and attachment names are indexed; attachment file contents are not bulk-downloaded.",
              "- Drive text is indexed for Google Docs/Sheets/Slides/Drawings, text, PDF, DOCX, XLSX and PPTX within 30 MB.",
              "- Images, video, audio, archives, scanned PDFs without embedded text, and files over 30 MB remain metadata-only.",
              "- Gmail filtering is keyword-based; Drive is not company-filtered.",
              "- The index is a reference cache. Current-state questions should still be checked against live Google.", ""]
    REPORT_PATH.write_text("\n".join(lines), encoding="utf-8")
    receipt("google_index_report_written", str(REPORT_PATH))


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--mode", choices=["gmail", "drive", "all", "report"], default="all")
    parser.add_argument("--max-gmail", type=int)
    parser.add_argument("--max-drive", type=int)
    parser.add_argument("--max-minutes", type=float, default=DEFAULT_MAX_MINUTES,
                        help="wall-clock ceiling for this run; 0 disables (not advised). "
                             "The run is resumable: already-indexed ids are skipped, so "
                             "the next run continues where this one stopped.")
    args = parser.parse_args()
    budget = Budget(args.max_minutes)
    con = db_open()
    gmail_stats = drive_stats = None
    try:
        if args.mode != "report":
            creds = google_auth.get_creds(interactive=False)
            if args.mode in {"gmail", "all"}:
                gmail = build("gmail", "v1", credentials=creds, cache_discovery=False)
                gmail_stats = gmail_index(con, gmail, args.max_gmail, budget)
            if args.mode in {"drive", "all"}:
                drive = build("drive", "v3", credentials=creds, cache_discovery=False)
                drive_stats = drive_index(con, drive, args.max_drive, budget)
        report(con, gmail_stats, drive_stats)
        incomplete = any((s or {}).get("stopped_early") for s in (gmail_stats, drive_stats))
        summary = {
            "status": "stopped_at_time_limit" if incomplete else "complete",
            "database": str(DB_PATH), "report": str(REPORT_PATH),
            "gmail_run": gmail_stats, "drive_run": drive_stats,
        }
        if incomplete:
            summary["resume"] = ("Time limit reached. Re-run the same command to continue; "
                                 "already-indexed items are skipped.")
        receipt("google_index_run_finished", f"{DB_PATH}#{summary['status']}")
        print(json.dumps(summary, indent=2), flush=True)
        return 0
    finally:
        con.close()


if __name__ == "__main__":
    raise SystemExit(main())
