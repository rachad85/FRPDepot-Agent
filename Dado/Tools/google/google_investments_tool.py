#!/usr/bin/env python
"""FRP Depot Investments Workbook Approved Update Tool.

Commissioned by Rachad Homsi on 2026-07-26.

The only remote write in this module is Drive files.update on the exact workbook
My Drive/My Files/Rachad/Bussiness Folder/Investements.xlsx. It cannot create,
delete, rename, move, share, or change permissions. The only workbook operation
currently commissioned is adding a cash-profit row to the Pistavo Labs revenue
section after a 24-hour, full-SHA-256, single-use staged approval.
"""
from __future__ import annotations

import argparse
from datetime import date, datetime, timedelta, timezone
from decimal import Decimal, InvalidOperation
import hashlib
import io
import json
import os
import posixpath
from pathlib import Path
import re
import secrets
import sys
from typing import Any
from xml.sax.saxutils import escape as xml_escape
from xml.etree import ElementTree as ET
from xml.parsers import expat
import zipfile

from openpyxl import load_workbook
from openpyxl.utils.datetime import to_excel
from googleapiclient.http import MediaIoBaseDownload, MediaIoBaseUpload

import google_investments_auth as auth

TOOL_NAME = "FRP Depot Investments Workbook Approved Update Tool"
SCHEMA_VERSION = 1
ACTION = "pistavo_cash_profit_add"
APPROVAL_WORD = "APPROVED"
PLAN_LIFETIME_HOURS = 24
ROOT = Path(r"C:\FRPDepot")
PLAN_DIR = ROOT / "Dado" / "20_Working" / "investments_plans"
RECEIPTS = ROOT / "Dado" / "40_Logs" / "receipts.jsonl"
BACKUP_DIR = auth.VAULT / "backups"
WORKBOOK_NAME = "Investements.xlsx"
WORKBOOK_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
EXPECTED_FILE_ID = "10QT40t5oTYa8jKD2ykIW-6X-RcyPOs2G"
EXPECTED_PARENT_PATH = ("My Drive", "My Files", "Rachad", "Bussiness Folder")
EXPECTED_PARENT_IDS = (
    "0ACKbTL9Q6AISUk9PVA",
    "1y-WtVIKC0APKhFN_GjEC5wBRjsLODzhQ",
    "14JcHCth2XM1968eQn1hA4jGKquQqDxMx",
    "12C-CPb_1PWt-WHTQOd3PDLeJ_IV9zSdw",
)
SHEET_NAME = "Pistavo Labs"
REVENUE_TITLE_ROW = 14
REVENUE_HEADER_ROW = 15
REVENUE_FIRST_ROW = 16
REVENUE_LAST_ROW = 35
COMMENT = "Profits received from Pistavo Labs - CASH"
AMOUNT_RE = re.compile(r"^[0-9]{1,9}(?:\.[0-9]{1,2})?$")
SECRET_RE = re.compile(r"(?i)(?:ya29\.[A-Za-z0-9._-]+|access[_-]?token[=: ]+[^\s,}]+|refresh[_-]?token[=: ]+[^\s,}]+)")
NS_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
NS_DOC_REL = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
NS_PKG_REL = "http://schemas.openxmlformats.org/package/2006/relationships"


class InvestmentsError(RuntimeError):
    pass


def utc_now() -> datetime:
    return datetime.now(timezone.utc)


def canonical(value: Any) -> str:
    return json.dumps(value, sort_keys=True, separators=(",", ":"), ensure_ascii=False)


def digest_for(value: dict[str, Any]) -> str:
    return hashlib.sha256(canonical(value).encode("utf-8")).hexdigest()


def content_sha256(value: bytes) -> str:
    return hashlib.sha256(value).hexdigest()


def scrub(value: str) -> str:
    return SECRET_RE.sub("[REDACTED]", str(value))


def append_receipt(action: str, evidence: str) -> None:
    RECEIPTS.parent.mkdir(parents=True, exist_ok=True)
    row = {"ts": utc_now().isoformat(), "action": action, "evidence": evidence}
    with RECEIPTS.open("a", encoding="utf-8") as handle:
        handle.write(json.dumps(row, ensure_ascii=True) + "\n")


def clean_date(value: str) -> str:
    try:
        parsed = date.fromisoformat(str(value).strip())
    except ValueError as exc:
        raise InvestmentsError("date must be YYYY-MM-DD.") from exc
    return parsed.isoformat()


def clean_amount(value: str) -> str:
    text = str(value).strip()
    if not AMOUNT_RE.fullmatch(text):
        raise InvestmentsError("amount must be a positive CAD number with at most two decimals.")
    try:
        number = Decimal(text)
    except InvalidOperation as exc:
        raise InvestmentsError("amount is invalid.") from exc
    if number <= 0 or number > Decimal("100000000"):
        raise InvestmentsError("amount must be greater than 0 and no more than CAD 100,000,000.")
    return format(number.normalize(), "f")


def excel_amount(value: str) -> int | float:
    number = Decimal(value)
    return int(number) if number == number.to_integral_value() else float(number)


def clean_source(value: str) -> str:
    text = str(value or "").strip()
    if not text or len(text) > 500:
        raise InvestmentsError("source is required and must be at most 500 characters.")
    if any(ord(char) < 32 for char in text):
        raise InvestmentsError("source contains control characters.")
    return text


def _file_metadata(service, file_id: str) -> tuple[dict[str, Any], str]:
    item = service.files().get(
        fileId=file_id,
        supportsAllDrives=True,
        fields=(
            "id,title,mimeType,parents(id),md5Checksum,modifiedDate,version,"
            "editable,downloadUrl,etag"
        ),
    ).execute(num_retries=3)
    etag = str(item.get("etag") or "").strip()
    if not etag:
        raise InvestmentsError(
            "Google Drive v2 did not return an ETag; conditional update is unavailable."
        )
    return item, etag


def _parent_identity(service, item: dict[str, Any]) -> tuple[tuple[str, ...], tuple[str, ...]]:
    parents = list(item.get("parents") or [])
    if len(parents) != 1:
        raise InvestmentsError("The investments workbook must have exactly one parent folder.")
    names: list[str] = []
    identifiers: list[str] = []
    seen: set[str] = set()
    current = str((parents[0] or {}).get("id") or "")
    while current:
        if current in seen or len(names) > 12:
            raise InvestmentsError("Drive parent path is cyclic or unexpectedly deep.")
        seen.add(current)
        parent = service.files().get(
            fileId=current, supportsAllDrives=True, fields="id,title,parents(id)"
        ).execute(num_retries=3)
        names.append(str(parent.get("title") or ""))
        identifiers.append(str(parent.get("id") or ""))
        next_parents = list(parent.get("parents") or [])
        if len(next_parents) > 1:
            raise InvestmentsError("A workbook parent folder has multiple parents.")
        current = str((next_parents[0] or {}).get("id") or "") if next_parents else ""
    return tuple(reversed(identifiers)), tuple(reversed(names))


def _verify_exact_file(service, item: dict[str, Any]) -> dict[str, Any]:
    if (
        item.get("id") != EXPECTED_FILE_ID
        or item.get("title") != WORKBOOK_NAME
        or item.get("mimeType") != WORKBOOK_MIME
    ):
        raise InvestmentsError("The pinned Drive item is not the commissioned Excel workbook.")
    parent_ids, path = _parent_identity(service, item)
    if parent_ids != EXPECTED_PARENT_IDS or path != EXPECTED_PARENT_PATH:
        raise InvestmentsError(
            "The investments workbook is not at the commissioned Drive path: "
            + " / ".join(EXPECTED_PARENT_PATH)
        )
    if not item.get("downloadUrl"):
        raise InvestmentsError("The workbook cannot be downloaded for verification.")
    return {
        **item,
        "verified_parent_ids": list(parent_ids),
        "verified_parent_path": list(path),
    }


def resolve_workbook(service) -> tuple[dict[str, Any], str]:
    item, etag = _file_metadata(service, EXPECTED_FILE_ID)
    item = _verify_exact_file(service, item)
    if not item.get("editable"):
        raise InvestmentsError("Google reports that this account cannot edit the workbook.")
    return item, etag


def download_bytes(service, file_id: str) -> bytes:
    target = io.BytesIO()
    request = service.files().get_media(fileId=file_id, supportsAllDrives=True)
    downloader = MediaIoBaseDownload(target, request)
    done = False
    while not done:
        _, done = downloader.next_chunk(num_retries=3)
    return target.getvalue()


def _formula_range(formula: Any) -> tuple[int, int]:
    match = re.fullmatch(r"\s*=\s*SUM\s*\(\s*B(\d+)\s*:\s*B(\d+)\s*\)\s*", str(formula or ""), re.I)
    if not match:
        raise InvestmentsError("Pistavo Labs revenue total formula changed; stopped.")
    return int(match.group(1)), int(match.group(2))


def _package(data: bytes) -> tuple[list[zipfile.ZipInfo], dict[str, bytes]]:
    try:
        with zipfile.ZipFile(io.BytesIO(data), "r") as archive:
            infos = archive.infolist()
            parts = {info.filename: archive.read(info.filename) for info in infos}
    except (zipfile.BadZipFile, KeyError) as exc:
        raise InvestmentsError("The investments workbook is not a valid XLSX package.") from exc
    if "xl/workbook.xml" not in parts or "xl/_rels/workbook.xml.rels" not in parts:
        raise InvestmentsError("The XLSX workbook manifest is incomplete.")
    if len(infos) != len(parts):
        raise InvestmentsError("The XLSX package contains duplicate member names.")
    return infos, parts


def _archive_comment(data: bytes) -> bytes:
    try:
        with zipfile.ZipFile(io.BytesIO(data), "r") as archive:
            return bytes(archive.comment)
    except zipfile.BadZipFile as exc:
        raise InvestmentsError("The investments workbook is not a valid XLSX package.") from exc


def _worksheet_path(parts: dict[str, bytes]) -> str:
    workbook = ET.fromstring(parts["xl/workbook.xml"])
    relationship_id = ""
    for sheet in workbook.findall(f".//{{{NS_MAIN}}}sheet"):
        if sheet.get("name") == SHEET_NAME:
            relationship_id = str(sheet.get(f"{{{NS_DOC_REL}}}id") or "")
            break
    if not relationship_id:
        raise InvestmentsError("Pistavo Labs worksheet relationship is missing.")
    relationships = ET.fromstring(parts["xl/_rels/workbook.xml.rels"])
    target = ""
    for relation in relationships.findall(f".//{{{NS_PKG_REL}}}Relationship"):
        if relation.get("Id") == relationship_id:
            target = str(relation.get("Target") or "")
            break
    if not target:
        raise InvestmentsError("Pistavo Labs worksheet target is missing.")
    normalized = (
        posixpath.normpath(target.lstrip("/"))
        if target.startswith("/")
        else posixpath.normpath(posixpath.join("xl", target))
    )
    if normalized not in parts:
        raise InvestmentsError("Pistavo Labs worksheet XML is missing from the XLSX package.")
    return normalized


def _row_cells(sheet_xml: bytes) -> dict[int, set[str]]:
    root = ET.fromstring(sheet_xml)
    result: dict[int, set[str]] = {}
    for row in root.findall(f".//{{{NS_MAIN}}}row"):
        try:
            number = int(str(row.get("r") or "0"))
        except ValueError:
            continue
        result[number] = {
            str(cell.get("r") or "") for cell in row.findall(f"{{{NS_MAIN}}}c")
        }
    return result


def _semantic_row_cells(sheet_xml: bytes) -> dict[int, set[str]]:
    """Return only cells containing a value, formula, or inline/shared string."""
    root = ET.fromstring(sheet_xml)
    result: dict[int, set[str]] = {}
    for row in root.findall(f".//{{{NS_MAIN}}}row"):
        try:
            number = int(str(row.get("r") or "0"))
        except ValueError:
            continue
        occupied: set[str] = set()
        for cell in row.findall(f"{{{NS_MAIN}}}c"):
            formula = cell.find(f"{{{NS_MAIN}}}f")
            inline = cell.find(f"{{{NS_MAIN}}}is")
            value = cell.find(f"{{{NS_MAIN}}}v")
            if formula is not None or inline is not None or (
                value is not None and value.text not in (None, "")
            ):
                occupied.add(str(cell.get("r") or ""))
        result[number] = occupied
    return result


def _style_id(sheet_xml: bytes, coordinate: str) -> str:
    root = ET.fromstring(sheet_xml)
    for cell in root.findall(f".//{{{NS_MAIN}}}c"):
        if cell.get("r") == coordinate:
            return str(cell.get("s") or "")
    return ""


def _actual_calcpr_start_span(workbook_xml: bytes) -> tuple[int, int]:
    direct_spans: list[tuple[int, int]] = []
    calcpr_count = 0
    stack: list[str] = []
    workbook_name = f"{NS_MAIN}}}workbook"
    calcpr_name = f"{NS_MAIN}}}calcPr"
    parser = expat.ParserCreate(namespace_separator="}")

    def start_element(name: str, _attributes: dict[str, str]) -> None:
        nonlocal calcpr_count
        parent = stack[-1] if stack else ""
        if name == calcpr_name:
            calcpr_count += 1
            if parent == workbook_name and len(stack) == 1:
                start = parser.CurrentByteIndex
                quote: int | None = None
                index = start
                while index < len(workbook_xml):
                    value = workbook_xml[index]
                    if quote is None and value in (34, 39):
                        quote = value
                    elif quote == value:
                        quote = None
                    elif quote is None and value == 62:
                        direct_spans.append((start, index + 1))
                        break
                    index += 1
                else:
                    raise InvestmentsError("Workbook calcPr start tag is incomplete.")
        stack.append(name)

    def end_element(_name: str) -> None:
        if stack:
            stack.pop()

    parser.StartElementHandler = start_element
    parser.EndElementHandler = end_element
    try:
        parser.Parse(workbook_xml, True)
    except expat.ExpatError as exc:
        raise InvestmentsError("Workbook XML is invalid.") from exc
    if calcpr_count != 1 or len(direct_spans) != 1:
        raise InvestmentsError(
            "Workbook must contain exactly one calcPr as a direct workbook child."
        )
    return direct_spans[0]


def _validate_recalculation(workbook_xml: bytes) -> None:
    try:
        root = ET.fromstring(workbook_xml)
    except ET.ParseError as exc:
        raise InvestmentsError("Workbook XML is invalid after recalculation patch.") from exc
    workbook_tag = f"{{{NS_MAIN}}}workbook"
    calcpr_tag = f"{{{NS_MAIN}}}calcPr"
    if root.tag != workbook_tag:
        raise InvestmentsError("Workbook XML root is not the SpreadsheetML workbook element.")
    descendants = root.findall(f".//{calcpr_tag}")
    direct = root.findall(f"./{calcpr_tag}")
    if len(descendants) != 1 or len(direct) != 1:
        raise InvestmentsError(
            "Workbook must contain exactly one calcPr as a direct workbook child."
        )
    element = direct[0]
    if (
        element.get("calcMode") != "auto"
        or element.get("fullCalcOnLoad") != "1"
        or element.get("forceFullCalc") != "1"
    ):
        raise InvestmentsError("Workbook recalculation settings are not enforced.")


def _force_recalculation(workbook_xml: bytes) -> bytes:
    """Patch the real calcPr element; comments and text cannot be selected."""
    start, end = _actual_calcpr_start_span(workbook_xml)
    tag = workbook_xml[start:end]
    closing = b"/>" if tag.endswith(b"/>") else b">"
    body = tag[:-len(closing)]
    for name, value in (
        (b"calcMode", b"auto"),
        (b"fullCalcOnLoad", b"1"),
        (b"forceFullCalc", b"1"),
    ):
        body = re.sub(rb"\s+" + name + rb'="[^"]*"', b"", body)
        body += b" " + name + b'="' + value + b'"'
    updated = workbook_xml[:start] + body + closing + workbook_xml[end:]
    _validate_recalculation(updated)
    return updated


def _replace_existing_cell(sheet_xml: bytes, coordinate: str, cell_type: str,
                           inner: bytes) -> tuple[bytes, bool]:
    coord = coordinate.encode("ascii")
    full = re.compile(
        rb'(<c\b(?=[^>]*\br="' + coord + rb'")[^>]*>)(.*?)(</c>)', re.DOTALL
    )
    match = full.search(sheet_xml)
    if match:
        opening = re.sub(rb'\s+t="[^"]*"', b"", match.group(1))
        opening = opening[:-1] + b' t="' + cell_type.encode("ascii") + b'">'
        replacement = opening + inner + match.group(3)
        return sheet_xml[:match.start()] + replacement + sheet_xml[match.end():], True
    empty = re.compile(rb'<c\b(?=[^>]*\br="' + coord + rb'")[^>]*/>')
    match = empty.search(sheet_xml)
    if match:
        opening = re.sub(rb'\s+t="[^"]*"', b"", match.group(0)[:-2])
        replacement = opening + b' t="' + cell_type.encode("ascii") + b'">' + inner + b"</c>"
        return sheet_xml[:match.start()] + replacement + sheet_xml[match.end():], True
    return sheet_xml, False


def _patch_sheet_xml(sheet_xml: bytes, target: int, entry: dict[str, Any], epoch) -> bytes:
    row_map = _row_cells(sheet_xml)
    semantic = _semantic_row_cells(sheet_xml).get(target, set())
    if semantic:
        raise InvestmentsError(
            f"Target row {target} contains data or formulas ({', '.join(sorted(semantic))}); stopped."
        )
    serial = to_excel(date.fromisoformat(str(entry["date"])), epoch=epoch)
    serial_text = str(int(serial)) if float(serial).is_integer() else format(serial, ".15g")
    amount_text = str(entry["amount_cad"])
    values = {
        f"A{target}": ("n", b"<v>" + serial_text.encode("ascii") + b"</v>"),
        f"B{target}": ("n", b"<v>" + amount_text.encode("ascii") + b"</v>"),
        f"C{target}": (
            "inlineStr", b"<is><t>" + xml_escape(COMMENT).encode("utf-8") + b"</t></is>"
        ),
    }
    existing = row_map.get(target, set())
    required = set(values)
    if required.issubset(existing):
        patched = sheet_xml
        for coordinate in (f"A{target}", f"B{target}", f"C{target}"):
            cell_type, inner = values[coordinate]
            patched, found = _replace_existing_cell(patched, coordinate, cell_type, inner)
            if not found:
                raise InvestmentsError(f"Blank cell {coordinate} could not be patched safely.")
        return patched
    if existing:
        raise InvestmentsError(
            f"Target row {target} has a partial cell layout; stopped rather than restructure it."
        )

    style_row = next(
        (row for row in range(target - 1, REVENUE_FIRST_ROW - 1, -1)
         if all(f"{column}{row}" in row_map.get(row, set()) for column in "ABC")),
        None,
    )
    if style_row is None:
        raise InvestmentsError("No prior Pistavo Labs revenue row is available for A:C formatting.")

    def style(column: str) -> bytes:
        identifier = _style_id(sheet_xml, f"{column}{style_row}")
        return (b' s="' + identifier.encode("ascii") + b'"') if identifier else b""

    row_number = str(target).encode("ascii")
    cells = (
        b'<c r="A' + row_number + b'"' + style("A")
        + b' t="n"><v>' + serial_text.encode("ascii") + b'</v></c>'
        + b'<c r="B' + row_number + b'"' + style("B")
        + b' t="n"><v>' + amount_text.encode("ascii") + b'</v></c>'
        + b'<c r="C' + row_number + b'"' + style("C")
        + b' t="inlineStr"><is><t>' + xml_escape(COMMENT).encode("utf-8") + b'</t></is></c>'
    )
    full_row = re.compile(
        rb'(<row\b(?=[^>]*\br="' + row_number + rb'")[^>]*>)(.*?)(</row>)', re.DOTALL
    )
    match = full_row.search(sheet_xml)
    if match:
        return sheet_xml[:match.start(2)] + match.group(2) + cells + sheet_xml[match.end(2):]
    empty_row = re.compile(rb'<row\b(?=[^>]*\br="' + row_number + rb'")[^>]*/>')
    match = empty_row.search(sheet_xml)
    if match:
        opening = match.group(0)[:-2] + b">"
        replacement = opening + cells + b"</row>"
        return sheet_xml[:match.start()] + replacement + sheet_xml[match.end():]
    insertion = b'<row r="' + row_number + b'">' + cells + b"</row>"
    for row_match in re.finditer(rb'<row\b[^>]*\br="(\d+)"[^>]*>', sheet_xml):
        if int(row_match.group(1)) > target:
            return sheet_xml[:row_match.start()] + insertion + sheet_xml[row_match.start():]
    end_marker = b"</sheetData>"
    if end_marker not in sheet_xml:
        raise InvestmentsError("Worksheet sheetData closing tag is missing.")
    return sheet_xml.replace(end_marker, insertion + end_marker, 1)


def _repack(infos: list[zipfile.ZipInfo], parts: dict[str, bytes], archive_comment: bytes) -> bytes:
    output = io.BytesIO()
    with zipfile.ZipFile(output, "w") as archive:
        for info in infos:
            archive.writestr(info, parts[info.filename])
        archive.comment = archive_comment
    return output.getvalue()


def _verify_preserved_parts(original: bytes, updated: bytes, sheet_path: str) -> None:
    _, before = _package(original)
    _, after = _package(updated)
    if set(before) != set(after):
        raise InvestmentsError("XLSX package parts changed unexpectedly.")
    if _archive_comment(original) != _archive_comment(updated):
        raise InvestmentsError("XLSX archive comment changed unexpectedly.")
    allowed = {sheet_path, "xl/workbook.xml"}
    changed = {name for name in before if before[name] != after[name]}
    if not changed or not changed.issubset(allowed) or sheet_path not in changed:
        raise InvestmentsError(
            "The XLSX patch changed files outside the approved worksheet/calculation metadata."
        )


def inspect_workbook(data: bytes, entry_date: str, amount_cad: str) -> dict[str, Any]:
    try:
        workbook = load_workbook(io.BytesIO(data), data_only=False, read_only=True)
    except Exception as exc:
        raise InvestmentsError("The live investments workbook cannot be opened as XLSX.") from exc
    if SHEET_NAME not in workbook.sheetnames:
        raise InvestmentsError("Pistavo Labs sheet is missing.")
    sheet = workbook[SHEET_NAME]
    if sheet.cell(REVENUE_TITLE_ROW, 1).value != "Revenues":
        raise InvestmentsError("Pistavo Labs revenue section moved or changed.")
    if str(sheet.cell(REVENUE_HEADER_ROW, 1).value or "").strip().casefold() != "date":
        raise InvestmentsError("Pistavo Labs revenue date header changed.")
    if str(sheet.cell(REVENUE_HEADER_ROW, 2).value or "").strip().casefold() != "cad":
        raise InvestmentsError("Pistavo Labs revenue currency is no longer CAD.")
    start, end = _formula_range(sheet.cell(REVENUE_TITLE_ROW, 2).value)
    if start != REVENUE_FIRST_ROW or end != REVENUE_LAST_ROW:
        raise InvestmentsError("Pistavo Labs revenue formula range changed.")
    wanted_date = date.fromisoformat(entry_date)
    wanted_amount = excel_amount(amount_cad)
    for row in range(start, end + 1):
        existing_date = sheet.cell(row, 1).value
        if isinstance(existing_date, datetime):
            existing_date = existing_date.date()
        existing_amount = sheet.cell(row, 2).value
        existing_comment = str(sheet.cell(row, 3).value or "").casefold()
        if existing_date == wanted_date and existing_amount == wanted_amount:
            if "pistavo" in existing_comment and "cash" in existing_comment:
                raise InvestmentsError(f"The requested entry already appears at row {row}; stopped.")
    _, parts = _package(data)
    sheet_path = _worksheet_path(parts)
    rows = _semantic_row_cells(parts[sheet_path])
    target = next((row for row in range(start, end + 1) if not rows.get(row)), None)
    if target is None:
        raise InvestmentsError("No blank Pistavo Labs revenue row remains inside the total formula.")
    return {
        "sheet": SHEET_NAME,
        "section": "Revenues",
        "row": target,
        "date": entry_date,
        "amount_cad": amount_cad,
        "comment": COMMENT,
        "total_formula": str(sheet.cell(REVENUE_TITLE_ROW, 2).value),
    }


def apply_entry(data: bytes, entry: dict[str, Any]) -> bytes:
    expected = inspect_workbook(data, str(entry["date"]), str(entry["amount_cad"]))
    if expected != entry:
        raise InvestmentsError("The approved target row no longer matches the workbook.")
    workbook = load_workbook(io.BytesIO(data), data_only=False, read_only=True)
    infos, parts = _package(data)
    sheet_path = _worksheet_path(parts)
    parts[sheet_path] = _patch_sheet_xml(
        parts[sheet_path], int(entry["row"]), entry, workbook.epoch
    )
    parts["xl/workbook.xml"] = _force_recalculation(parts["xl/workbook.xml"])
    value = _repack(infos, parts, _archive_comment(data))
    _verify_preserved_parts(data, value, sheet_path)
    verify_entry(value, entry)
    return value


def verify_entry(data: bytes, entry: dict[str, Any]) -> None:
    workbook = load_workbook(io.BytesIO(data), data_only=False, read_only=True)
    sheet = workbook[SHEET_NAME]
    row = int(entry["row"])
    actual_date = sheet.cell(row, 1).value
    if isinstance(actual_date, datetime):
        actual_date = actual_date.date()
    start, end = _formula_range(sheet.cell(REVENUE_TITLE_ROW, 2).value)
    if not (
        actual_date == date.fromisoformat(str(entry["date"]))
        and sheet.cell(row, 2).value == excel_amount(str(entry["amount_cad"]))
        and sheet.cell(row, 3).value == COMMENT
        and start <= row <= end
    ):
        raise InvestmentsError("Workbook readback does not match the approved entry.")
    _, parts = _package(data)
    _validate_recalculation(parts["xl/workbook.xml"])
    sheet_xml = parts[_worksheet_path(parts)]
    physical = _row_cells(sheet_xml).get(row, set())
    semantic = _semantic_row_cells(sheet_xml).get(row, set())
    required = {f"A{row}", f"B{row}", f"C{row}"}
    if semantic != required or not required.issubset(physical):
        raise InvestmentsError("Approved row contains data outside A:C after patching.")


def approval_phrase(digest: str) -> str:
    # The full digest remains validated internally; Rachad approved a plain,
    # one-word confirmation so he never has to copy checksum characters.
    return APPROVAL_WORD


def lock_path(plan_digest: str) -> Path:
    if not re.fullmatch(r"[0-9a-f]{64}", str(plan_digest)):
        raise InvestmentsError("Plan digest is invalid for replay locking.")
    return PLAN_DIR / ".commit-locks" / f"{plan_digest}.json"


def write_lock(path: Path, value: dict[str, Any], *, exclusive: bool = False) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    flags = os.O_WRONLY | os.O_CREAT | (os.O_EXCL if exclusive else os.O_TRUNC)
    try:
        descriptor = os.open(str(path), flags, 0o600)
    except FileExistsError as exc:
        raise InvestmentsError("This plan has already entered commit and cannot be replayed.") from exc
    with os.fdopen(descriptor, "w", encoding="utf-8", newline="\n") as handle:
        handle.write(json.dumps(value, ensure_ascii=True, indent=2) + "\n")


def command_connect(_args: argparse.Namespace) -> None:
    service = auth.drive_service(interactive=True, force=True)
    item, _ = resolve_workbook(service)
    print(json.dumps({
        "status": "CONNECTED_AND_VERIFIED",
        "account": auth.EXPECTED_ACCOUNT,
        "permission": "Full Drive OAuth; tool write allowlist is exact workbook only",
        "workbook": WORKBOOK_NAME,
        "path": " / ".join(EXPECTED_PARENT_PATH),
        "can_edit": bool(item.get("editable")),
        "token_vault": str(auth.VAULT),
    }, indent=2))


def command_check(_args: argparse.Namespace) -> None:
    service = auth.drive_service()
    item, _ = resolve_workbook(service)
    print(json.dumps({
        "status": "ACCESS_VERIFIED_READ_ONLY_CHECK",
        "account": auth.EXPECTED_ACCOUNT,
        "workbook": WORKBOOK_NAME,
        "path": " / ".join(EXPECTED_PARENT_PATH),
        "can_edit": bool(item.get("editable")),
        "remote_write_performed": False,
    }, indent=2))


def command_stage(args: argparse.Namespace) -> None:
    entry_date = clean_date(args.date)
    amount_cad = clean_amount(args.amount)
    source = clean_source(args.source)
    service = auth.drive_service()
    item, etag = resolve_workbook(service)
    current = download_bytes(service, str(item["id"]))
    entry = inspect_workbook(current, entry_date, amount_cad)
    # Exercise the exact package-preserving edit locally before staging.
    apply_entry(current, entry)
    created = utc_now()
    core = {
        "schema_version": SCHEMA_VERSION,
        "tool": TOOL_NAME,
        "action": ACTION,
        "created_utc": created.isoformat(),
        "expires_utc": (created + timedelta(hours=PLAN_LIFETIME_HOURS)).isoformat(),
        "nonce": secrets.token_hex(16),
        "file": {
            "id": str(item["id"]),
            "name": WORKBOOK_NAME,
            "parent_ids": list(EXPECTED_PARENT_IDS),
            "parent_path": list(EXPECTED_PARENT_PATH),
            "mime_type": WORKBOOK_MIME,
            "etag": etag,
            "md5_checksum": str(item.get("md5Checksum") or ""),
            "modified_time": str(item.get("modifiedDate") or ""),
            "version": str(item.get("version") or ""),
            "content_sha256": content_sha256(current),
        },
        "entry": entry,
        "source": source,
    }
    digest = digest_for(core)
    plan = {**core, "sha256": digest}
    PLAN_DIR.mkdir(parents=True, exist_ok=True)
    stamp = created.strftime("%Y%m%dT%H%M%SZ")
    path = PLAN_DIR / f"{stamp}_{ACTION}_{digest[:16]}.json"
    path.write_text(json.dumps(plan, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    append_receipt("investments_update_plan_staged", str(path))
    print(json.dumps({
        "status": "STAGED_NOT_COMMITTED",
        "plan": str(path),
        "expires_utc": plan["expires_utc"],
        "workbook": WORKBOOK_NAME,
        "path": " / ".join(EXPECTED_PARENT_PATH),
        "entry": entry,
        "source": source,
        "approval": approval_phrase(digest),
        "remote_write_performed": False,
    }, indent=2, ensure_ascii=False))


def read_json(path: Path) -> dict[str, Any]:
    try:
        value = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        raise InvestmentsError("Plan is unreadable.") from exc
    if not isinstance(value, dict):
        raise InvestmentsError("Plan must contain one JSON object.")
    return value


def load_plan(path: Path) -> dict[str, Any]:
    plan = read_json(path)
    saved = str(plan.pop("sha256", ""))
    if not saved or not secrets.compare_digest(saved, digest_for(plan)):
        raise InvestmentsError("Plan hash check failed. The plan changed after review.")
    expected_keys = {
        "schema_version", "tool", "action", "created_utc", "expires_utc",
        "nonce", "file", "entry", "source",
    }
    if set(plan) != expected_keys:
        raise InvestmentsError("Plan fields do not match the closed schema.")
    if plan.get("schema_version") != SCHEMA_VERSION or plan.get("tool") != TOOL_NAME or plan.get("action") != ACTION:
        raise InvestmentsError("Plan schema, tool, or action is invalid.")
    try:
        created = datetime.fromisoformat(str(plan["created_utc"]))
        expires = datetime.fromisoformat(str(plan["expires_utc"]))
    except (KeyError, ValueError, TypeError) as exc:
        raise InvestmentsError("Plan creation or expiry time is invalid.") from exc
    if created.tzinfo is None or expires.tzinfo is None:
        raise InvestmentsError("Plan times must include a timezone.")
    if expires - created != timedelta(hours=PLAN_LIFETIME_HOURS):
        raise InvestmentsError("Plan lifetime must be exactly 24 hours.")
    now = utc_now()
    if created > now + timedelta(minutes=5):
        raise InvestmentsError("Plan creation time is in the future.")
    if now >= expires:
        raise InvestmentsError("Plan expired. Stage a new plan.")
    if not re.fullmatch(r"[0-9a-f]{32}", str(plan.get("nonce") or "")):
        raise InvestmentsError("Plan nonce is invalid.")
    file_info = plan.get("file")
    if not isinstance(file_info, dict):
        raise InvestmentsError("Plan file identity is invalid.")
    expected_file_keys = {
        "id", "name", "parent_ids", "parent_path", "mime_type", "etag",
        "md5_checksum", "modified_time", "version", "content_sha256",
    }
    if (
        set(file_info) != expected_file_keys
        or file_info.get("id") != EXPECTED_FILE_ID
        or file_info.get("name") != WORKBOOK_NAME
        or file_info.get("parent_ids") != list(EXPECTED_PARENT_IDS)
        or file_info.get("parent_path") != list(EXPECTED_PARENT_PATH)
        or file_info.get("mime_type") != WORKBOOK_MIME
        or not str(file_info.get("etag") or "")
        or not re.fullmatch(r"[0-9a-f]{64}", str(file_info.get("content_sha256") or ""))
    ):
        raise InvestmentsError("Plan does not target the commissioned workbook.")
    entry = plan.get("entry")
    if not isinstance(entry, dict):
        raise InvestmentsError("Plan entry is invalid.")
    rebuilt = {
        "sheet": SHEET_NAME,
        "section": "Revenues",
        "row": int(entry.get("row") or 0),
        "date": clean_date(str(entry.get("date") or "")),
        "amount_cad": clean_amount(str(entry.get("amount_cad") or "")),
        "comment": COMMENT,
        "total_formula": str(entry.get("total_formula") or ""),
    }
    if rebuilt != entry or not (REVENUE_FIRST_ROW <= rebuilt["row"] <= REVENUE_LAST_ROW):
        raise InvestmentsError("Plan entry failed the current allowlist validation.")
    if clean_source(str(plan.get("source") or "")) != plan.get("source"):
        raise InvestmentsError("Plan source failed validation.")
    plan["sha256"] = saved
    return plan


def _same_live_state(item: dict[str, Any], planned: dict[str, Any]) -> bool:
    return all(
        str(item.get(live) or "") == str(planned.get(saved) or "")
        for live, saved in (
            ("id", "id"),
            ("md5Checksum", "md5_checksum"),
            ("modifiedDate", "modified_time"),
            ("version", "version"),
        )
    )


def command_commit(args: argparse.Namespace) -> None:
    plan_path = Path(args.plan).resolve()
    if PLAN_DIR.resolve() not in plan_path.parents:
        raise InvestmentsError("Plan must be inside Dado's investments plan folder.")
    plan = load_plan(plan_path)
    expected = approval_phrase(str(plan["sha256"]))
    if not secrets.compare_digest(str(args.approval).strip().casefold(), expected.casefold()):
        raise InvestmentsError("Rachad must reply with the one-word approval: APPROVED.")
    lock = lock_path(str(plan["sha256"]))
    if lock.exists():
        raise InvestmentsError("This plan has already entered commit and cannot be replayed.")
    service = auth.drive_service()
    item, etag = _file_metadata(service, EXPECTED_FILE_ID)
    item = _verify_exact_file(service, item)
    if not item.get("editable"):
        raise InvestmentsError("Google no longer permits editing this workbook.")
    if not _same_live_state(item, plan["file"]) or etag != plan["file"]["etag"]:
        raise InvestmentsError("The live workbook changed after review. Stage a new plan.")
    current = download_bytes(service, str(item["id"]))
    if content_sha256(current) != plan["file"]["content_sha256"]:
        raise InvestmentsError("The live workbook bytes changed after review. Stage a new plan.")
    updated = apply_entry(current, plan["entry"])
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    backup = BACKUP_DIR / (
        utc_now().strftime("%Y%m%dT%H%M%SZ_") + str(plan["file"]["md5_checksum"]) + "_Investements.xlsx"
    )
    backup.write_bytes(current)
    write_lock(lock, {
        "plan_sha256": plan["sha256"],
        "status": "in_flight",
        "started_utc": utc_now().isoformat(),
        "backup": str(backup),
    }, exclusive=True)
    try:
        # Re-check identity/state immediately before the one remote write, then
        # bind that write to Google's current ETag. If anything changes after
        # this check, Google must reject the upload with HTTP 412.
        preflight, preflight_etag = _file_metadata(service, EXPECTED_FILE_ID)
        preflight = _verify_exact_file(service, preflight)
        if not _same_live_state(preflight, plan["file"]) or preflight_etag != plan["file"]["etag"]:
            raise InvestmentsError("The live workbook changed immediately before upload.")
        if utc_now() >= datetime.fromisoformat(str(plan["expires_utc"])):
            raise InvestmentsError("Plan expired immediately before upload.")
        media = MediaIoBaseUpload(io.BytesIO(updated), mimetype=WORKBOOK_MIME, resumable=False)
        update_request = service.files().update(
            fileId=str(item["id"]),
            media_body=media,
            supportsAllDrives=True,
            fields="id,title,mimeType,md5Checksum,modifiedDate,version,etag",
        )
        update_request.headers["If-Match"] = preflight_etag
        update_request.execute(num_retries=0)
        readback = download_bytes(service, str(item["id"]))
        verify_entry(readback, plan["entry"])
        after, _ = _file_metadata(service, EXPECTED_FILE_ID)
        after = _verify_exact_file(service, after)
        if after.get("title") != WORKBOOK_NAME or content_sha256(readback) != content_sha256(updated):
            raise InvestmentsError("Live Drive readback does not match the approved workbook bytes.")
    except Exception as exc:
        write_lock(lock, {
            "plan_sha256": plan["sha256"],
            "status": "indeterminate",
            "updated_utc": utc_now().isoformat(),
            "backup": str(backup),
            "reason": scrub(str(exc)),
        })
        append_receipt(
            "investments_update_indeterminate_no_retry",
            f"plan={plan_path}; sha256={plan['sha256']}; backup={backup}",
        )
        raise InvestmentsError(
            "The Drive update failed or could not be verified. The plan is locked and will not retry. "
            "Use Drive version history and the local backup to reconcile it."
        ) from exc
    write_lock(lock, {
        "plan_sha256": plan["sha256"],
        "status": "committed_verified",
        "updated_utc": utc_now().isoformat(),
        "backup": str(backup),
        "new_md5_checksum": str(after.get("md5Checksum") or ""),
        "new_version": str(after.get("version") or ""),
    })
    append_receipt(
        "investments_approved_update_committed_verified",
        f"plan={plan_path}; sha256={plan['sha256']}; workbook={WORKBOOK_NAME}; backup={backup}",
    )
    print(json.dumps({
        "status": "COMMITTED_AND_VERIFIED",
        "workbook": WORKBOOK_NAME,
        "path": " / ".join(EXPECTED_PARENT_PATH),
        "entry": plan["entry"],
        "plan_sha256": plan["sha256"],
        "replay_locked": True,
        "local_backup": str(backup),
    }, indent=2, ensure_ascii=False))


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=TOOL_NAME)
    commands = parser.add_subparsers(dest="command", required=True)
    connect = commands.add_parser("connect")
    connect.set_defaults(func=command_connect)
    check = commands.add_parser("check")
    check.set_defaults(func=command_check)
    stage = commands.add_parser("stage-pistavo-cash-profit")
    stage.add_argument("--date", required=True)
    stage.add_argument("--amount", required=True)
    stage.add_argument("--source", required=True)
    stage.set_defaults(func=command_stage)
    commit = commands.add_parser("commit")
    commit.add_argument("--plan", required=True)
    commit.add_argument("--approval", required=True)
    commit.set_defaults(func=command_commit)
    return parser


def main() -> int:
    args = build_parser().parse_args()
    try:
        args.func(args)
        return 0
    except (InvestmentsError, auth.InvestmentsAuthError, OSError, ValueError) as exc:
        print("ERROR: " + scrub(str(exc)), file=sys.stderr)
        return 1
    except Exception as exc:
        print("ERROR: Investments tool failed safely: " + scrub(str(exc)), file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
