from __future__ import annotations

from datetime import datetime, timedelta, timezone
import io
import json
from pathlib import Path
import tempfile
import unittest
import zipfile

from openpyxl import Workbook, load_workbook

import google_investments_auth as auth
import google_investments_tool as tool


class InvestmentsToolTests(unittest.TestCase):
    def workbook_bytes(self, *, duplicate: bool = False, occupied_d17: bool = False,
                       styled_blank_d17: bool = False, custom_part: bool = False,
                       archive_comment: bytes = b"") -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = tool.SHEET_NAME
        ws["A14"] = "Revenues"
        ws["B14"] = "=SUM(B16:B35)"
        ws["A15"] = "Date"
        ws["B15"] = "CAD"
        ws["C15"] = "Comments"
        ws["A16"] = datetime(2026, 7, 13)
        ws["B16"] = 10000
        ws["C16"] = "CASH"
        ws["A16"].number_format = "yyyy-mm-dd"
        if duplicate:
            ws["A17"] = datetime(2026, 7, 26)
            ws["B17"] = 10000
            ws["C17"] = tool.COMMENT
        if occupied_d17:
            ws["D17"] = "=1+1"
        if styled_blank_d17:
            ws["A17"].number_format = "yyyy-mm-dd"
            ws["B17"].number_format = "0.00"
            ws["C17"].number_format = "@"
            ws["D17"].number_format = "0.0000"
        target = io.BytesIO()
        wb.save(target)
        value = target.getvalue()
        if custom_part or archive_comment:
            output = io.BytesIO()
            with zipfile.ZipFile(io.BytesIO(value), "r") as source, zipfile.ZipFile(output, "w") as dest:
                for info in source.infolist():
                    dest.writestr(info, source.read(info.filename))
                if custom_part:
                    dest.writestr("customXml/preserved-test.xml", b"<preserve exactly='yes'/>")
                dest.comment = archive_comment
            value = output.getvalue()
        return value

    def test_scope_is_separate_full_drive_only(self):
        self.assertEqual(auth.DRIVE_SCOPE, "https://www.googleapis.com/auth/drive")
        source = Path(auth.__file__).read_text(encoding="utf-8")
        self.assertNotIn("https://www.googleapis.com/auth/gmail", source)
        self.assertNotIn("drive.readonly", source)

        class FakeCredentials:
            scopes = [auth.DRIVE_SCOPE]
            granted_scopes = [auth.DRIVE_SCOPE]
            refresh_token = "high-entropy-test-refresh"
            client_id = "test-client"

        exact = FakeCredentials()
        self.assertTrue(auth._exact_requested_scope(exact))
        self.assertTrue(auth._exact_new_grant(exact))
        exact.granted_scopes = [auth.DRIVE_SCOPE, "https://mail.google.com/"]
        self.assertFalse(auth._exact_new_grant(exact))

    def test_exact_workbook_and_path_are_fixed(self):
        self.assertEqual(tool.WORKBOOK_NAME, "Investements.xlsx")
        self.assertEqual(
            tool.EXPECTED_PARENT_PATH,
            ("My Drive", "My Files", "Rachad", "Bussiness Folder"),
        )
        self.assertEqual(tool.SHEET_NAME, "Pistavo Labs")
        self.assertRegex(tool.EXPECTED_FILE_ID, r"^[A-Za-z0-9_-]+$")
        self.assertEqual(len(tool.EXPECTED_PARENT_IDS), 4)

    def test_amount_and_date_validation(self):
        self.assertEqual(tool.clean_amount("10000"), "10000")
        self.assertEqual(tool.clean_amount("10.50"), "10.5")
        self.assertEqual(tool.clean_date("2026-07-26"), "2026-07-26")
        for bad in ("0", "-1", "1.234", "=10000", "nan", "100000001"):
            with self.assertRaises(tool.InvestmentsError):
                tool.clean_amount(bad)
        with self.assertRaises(tool.InvestmentsError):
            tool.clean_date("07/26/2026")

    def test_inspect_and_apply_exact_entry(self):
        original = self.workbook_bytes()
        entry = tool.inspect_workbook(original, "2026-07-26", "10000")
        self.assertEqual(entry["row"], 17)
        self.assertEqual(entry["comment"], tool.COMMENT)
        updated = tool.apply_entry(original, entry)
        tool.verify_entry(updated, entry)
        wb = load_workbook(io.BytesIO(updated), data_only=False, read_only=True)
        ws = wb[tool.SHEET_NAME]
        self.assertEqual(ws["A17"].value.date().isoformat(), "2026-07-26")
        self.assertEqual(ws["B17"].value, 10000)
        self.assertEqual(ws["C17"].value, tool.COMMENT)
        self.assertEqual(ws["B14"].value, "=SUM(B16:B35)")

    def test_duplicate_is_refused(self):
        with self.assertRaisesRegex(tool.InvestmentsError, "already appears"):
            tool.inspect_workbook(self.workbook_bytes(duplicate=True), "2026-07-26", "10000")

    def test_occupied_d_cell_is_not_overwritten_and_custom_part_survives(self):
        original = self.workbook_bytes(occupied_d17=True, custom_part=True)
        entry = tool.inspect_workbook(original, "2026-07-26", "10000")
        self.assertEqual(entry["row"], 18)
        updated = tool.apply_entry(original, entry)
        wb = load_workbook(io.BytesIO(updated), data_only=False, read_only=True)
        self.assertEqual(wb[tool.SHEET_NAME]["D17"].value, "=1+1")
        with zipfile.ZipFile(io.BytesIO(updated), "r") as archive:
            self.assertEqual(
                archive.read("customXml/preserved-test.xml"), b"<preserve exactly='yes'/>"
            )

    def test_styled_blank_d17_and_archive_comment_are_preserved(self):
        original = self.workbook_bytes(
            styled_blank_d17=True, archive_comment=b"FRP Depot preserved archive comment"
        )
        entry = tool.inspect_workbook(original, "2026-07-26", "10000")
        self.assertEqual(entry["row"], 17)
        updated = tool.apply_entry(original, entry)
        wb = load_workbook(io.BytesIO(updated), data_only=False, read_only=True)
        ws = wb[tool.SHEET_NAME]
        self.assertIsNone(ws["D17"].value)
        self.assertEqual(ws["D17"].number_format, "0.0000")
        self.assertEqual(tool._archive_comment(updated), b"FRP Depot preserved archive comment")

    def test_calcpr_in_comment_cannot_bypass_real_recalculation(self):
        original = self.workbook_bytes()
        infos, parts = tool._package(original)
        workbook_xml = parts["xl/workbook.xml"]
        start, end = tool._actual_calcpr_start_span(workbook_xml)
        real = workbook_xml[start:end]
        body = real[:-2]
        for name in (b"calcMode", b"fullCalcOnLoad", b"forceFullCalc"):
            body = tool.re.sub(rb"\s+" + name + rb'="[^"]*"', b"", body)
        real = body + b' calcMode="manual" fullCalcOnLoad="0" forceFullCalc="0"/>'
        trap = b'<!-- <calcPr calcMode="auto" fullCalcOnLoad="1" forceFullCalc="1"/> -->'
        parts["xl/workbook.xml"] = workbook_xml[:start] + trap + real + workbook_xml[end:]
        trapped = tool._repack(infos, parts, tool._archive_comment(original))
        entry = tool.inspect_workbook(trapped, "2026-07-26", "10000")
        updated = tool.apply_entry(trapped, entry)
        _, updated_parts = tool._package(updated)
        tool._validate_recalculation(updated_parts["xl/workbook.xml"])
        self.assertIn(trap, updated_parts["xl/workbook.xml"])

    def test_nested_calcpr_is_rejected(self):
        nested = (
            b'<workbook xmlns="' + tool.NS_MAIN.encode("ascii") + b'">'
            b'<extLst><ext uri="test"><calcPr calcMode="auto" '
            b'fullCalcOnLoad="1" forceFullCalc="1"/></ext></extLst></workbook>'
        )
        with self.assertRaisesRegex(tool.InvestmentsError, "direct workbook child"):
            tool._actual_calcpr_start_span(nested)
        with self.assertRaisesRegex(tool.InvestmentsError, "direct workbook child"):
            tool._validate_recalculation(nested)

    def test_changed_formula_is_refused(self):
        raw = self.workbook_bytes()
        wb = load_workbook(io.BytesIO(raw))
        wb[tool.SHEET_NAME]["B14"] = "=SUM(B16:B36)"
        out = io.BytesIO(); wb.save(out)
        with self.assertRaisesRegex(tool.InvestmentsError, "formula range changed"):
            tool.inspect_workbook(out.getvalue(), "2026-07-26", "10000")

    def valid_plan(self):
        created = tool.utc_now()
        core = {
            "schema_version": tool.SCHEMA_VERSION,
            "tool": tool.TOOL_NAME,
            "action": tool.ACTION,
            "created_utc": created.isoformat(),
            "expires_utc": (created + timedelta(hours=tool.PLAN_LIFETIME_HOURS)).isoformat(),
            "nonce": "a" * 32,
            "file": {
                "id": tool.EXPECTED_FILE_ID,
                "name": tool.WORKBOOK_NAME,
                "parent_ids": list(tool.EXPECTED_PARENT_IDS),
                "parent_path": list(tool.EXPECTED_PARENT_PATH),
                "mime_type": tool.WORKBOOK_MIME,
                "etag": '"fixed-etag"',
                "md5_checksum": "abc",
                "modified_time": "2026-07-26T00:00:00Z",
                "version": "1",
                "content_sha256": "0" * 64,
            },
            "entry": {
                "sheet": tool.SHEET_NAME,
                "section": "Revenues",
                "row": 17,
                "date": "2026-07-26",
                "amount_cad": "10000",
                "comment": tool.COMMENT,
                "total_formula": "=SUM(B16:B35)",
            },
            "source": "Rachad Telegram instruction",
        }
        return {**core, "sha256": tool.digest_for(core)}

    def test_plan_hash_and_tamper_detection(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "plan.json"
            plan = self.valid_plan()
            path.write_text(json.dumps(plan), encoding="utf-8")
            loaded = tool.load_plan(path)
            self.assertEqual(loaded["sha256"], plan["sha256"])
            plan["entry"]["amount_cad"] = "20000"
            path.write_text(json.dumps(plan), encoding="utf-8")
            with self.assertRaisesRegex(tool.InvestmentsError, "hash check failed"):
                tool.load_plan(path)

    def test_expired_plan_is_refused(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "plan.json"
            plan = self.valid_plan()
            created = datetime.now(timezone.utc) - timedelta(hours=25)
            plan["created_utc"] = created.isoformat()
            plan["expires_utc"] = (created + timedelta(hours=24)).isoformat()
            core = dict(plan); core.pop("sha256")
            plan["sha256"] = tool.digest_for(core)
            path.write_text(json.dumps(plan), encoding="utf-8")
            with self.assertRaisesRegex(tool.InvestmentsError, "expired"):
                tool.load_plan(path)

    def test_remote_write_surface_is_one_exact_files_update(self):
        source = Path(tool.__file__).read_text(encoding="utf-8")
        self.assertEqual(source.count("service.files().update("), 1)
        self.assertNotIn("service.files().create(", source)
        self.assertNotIn("service.files().delete(", source)
        self.assertNotIn("service.files().copy(", source)
        self.assertNotIn("service.permissions()", source)
        self.assertNotIn("service.revisions()", source)
        self.assertIn('update_request.headers["If-Match"]', source)

    def test_replay_lock_is_digest_keyed_not_filename_keyed(self):
        digest = "a" * 64
        old_dir = tool.PLAN_DIR
        with tempfile.TemporaryDirectory() as tmp:
            try:
                tool.PLAN_DIR = Path(tmp)
                one = tool.lock_path(digest)
                two = tool.lock_path(digest)
                self.assertEqual(one, two)
                tool.write_lock(one, {"status": "in_flight"}, exclusive=True)
                with self.assertRaisesRegex(tool.InvestmentsError, "cannot be replayed"):
                    tool.write_lock(two, {"status": "in_flight"}, exclusive=True)
            finally:
                tool.PLAN_DIR = old_dir

    def test_approval_is_one_plain_word_while_digest_stays_internal(self):
        digest = "f" * 64
        self.assertEqual(tool.approval_phrase(digest), "APPROVED")
        self.assertNotIn(digest, tool.approval_phrase(digest))


if __name__ == "__main__":
    unittest.main()
